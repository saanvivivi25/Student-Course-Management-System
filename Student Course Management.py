import sqlite3
from openpyxl import Workbook


# ================= DATABASE =================

def create_database():
    conn = sqlite3.connect("student_management.db")
    cursor = conn.cursor()

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS Students(
        student_id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_name TEXT NOT NULL,
        age INTEGER,
        email TEXT NOT NULL
    )
    """)

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS Courses(
        course_id INTEGER PRIMARY KEY AUTOINCREMENT,
        course_name TEXT NOT NULL,
        course_duration TEXT
    )
    """)

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS Enrollments(
        enrollment_id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id INTEGER,
        course_id INTEGER,
        marks REAL,
        FOREIGN KEY(student_id) REFERENCES Students(student_id),
        FOREIGN KEY(course_id) REFERENCES Courses(course_id)
    )
    """)

    conn.commit()
    conn.close()


# ================= STUDENTS =================

def add_student():
    conn = sqlite3.connect("student_management.db")
    cursor = conn.cursor()

    while True:
        name = input("Enter Student Name: ").strip()
        if name:
            break
        print("Name cannot be empty.")

    while True:
        try:
            age = int(input("Enter Age: "))
            break
        except ValueError:
            print("Age must be numeric.")

    while True:
        email = input("Enter Email: ").strip()
        if email:
            break
        print("Email cannot be empty.")

    cursor.execute("""
    INSERT INTO Students(student_name, age, email)
    VALUES (?, ?, ?)
    """, (name, age, email))

    conn.commit()
    conn.close()

    print("Student Added Successfully.")


def view_students():
    conn = sqlite3.connect("student_management.db")
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM Students")
    students = cursor.fetchall()

    if not students:
        print("No students found.")
    else:
        for student in students:
            print(student)

    conn.close()


# ================= COURSES =================

def add_course():
    conn = sqlite3.connect("student_management.db")
    cursor = conn.cursor()

    course_name = input("Enter Course Name: ")
    duration = input("Enter Course Duration: ")

    cursor.execute("""
    INSERT INTO Courses(course_name, course_duration)
    VALUES (?, ?)
    """, (course_name, duration))

    conn.commit()
    conn.close()

    print("Course Added Successfully.")


def view_courses():
    conn = sqlite3.connect("student_management.db")
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM Courses")
    courses = cursor.fetchall()

    if not courses:
        print("No courses found.")
    else:
        for course in courses:
            print(course)

    conn.close()


# ================= ENROLLMENT =================

def enroll_student():
    conn = sqlite3.connect("student_management.db")
    cursor = conn.cursor()

    view_students()
    student_id = int(input("Enter Student ID: "))

    view_courses()
    course_id = int(input("Enter Course ID: "))

    marks = float(input("Enter Marks: "))

    cursor.execute("""
    INSERT INTO Enrollments(student_id, course_id, marks)
    VALUES (?, ?, ?)
    """, (student_id, course_id, marks))

    conn.commit()
    conn.close()

    print("Student Enrolled Successfully.")


def view_enrollments():
    conn = sqlite3.connect("student_management.db")
    cursor = conn.cursor()

    cursor.execute("""
    SELECT
    Students.student_name,
    Courses.course_name,
    Enrollments.marks
    FROM Enrollments
    JOIN Students
    ON Students.student_id = Enrollments.student_id
    JOIN Courses
    ON Courses.course_id = Enrollments.course_id
    """)

    records = cursor.fetchall()

    if not records:
        print("No enrollments found.")
    else:
        for row in records:
            print(row)

    conn.close()


# ================= UPDATE =================

def update_marks():
    conn = sqlite3.connect("student_management.db")
    cursor = conn.cursor()

    enrollment_id = int(input("Enter Enrollment ID: "))
    marks = float(input("Enter New Marks: "))

    cursor.execute("""
    UPDATE Enrollments
    SET marks = ?
    WHERE enrollment_id = ?
    """, (marks, enrollment_id))

    conn.commit()
    conn.close()

    print("Marks Updated Successfully.")


# ================= DELETE =================

def delete_enrollment():
    conn = sqlite3.connect("student_management.db")
    cursor = conn.cursor()

    enrollment_id = int(input("Enter Enrollment ID: "))

    cursor.execute("""
    DELETE FROM Enrollments
    WHERE enrollment_id = ?
    """, (enrollment_id,))

    conn.commit()
    conn.close()

    print("Enrollment Deleted Successfully.")


# ================= TOP STUDENT =================

def top_student():
    conn = sqlite3.connect("student_management.db")
    cursor = conn.cursor()

    cursor.execute("""
    SELECT
    Students.student_name,
    AVG(Enrollments.marks)
    FROM Students
    JOIN Enrollments
    ON Students.student_id = Enrollments.student_id
    GROUP BY Students.student_id
    """)

    records = cursor.fetchall()

    highest_average = -1
    top_name = ""

    for name, avg_marks in records:
        if avg_marks > highest_average:
            highest_average = avg_marks
            top_name = name

    if top_name:
        print("Top Student:", top_name)
        print("Average Marks:", round(highest_average, 2))
    else:
        print("No enrollments found.")

    conn.close()


# ================= COURSE AVERAGE =================

def course_average():
    conn = sqlite3.connect("student_management.db")
    cursor = conn.cursor()

    cursor.execute("""
    SELECT
    Courses.course_name,
    AVG(Enrollments.marks)
    FROM Courses
    JOIN Enrollments
    ON Courses.course_id = Enrollments.course_id
    GROUP BY Courses.course_id
    """)

    records = cursor.fetchall()

    if not records:
        print("No records found.")
    else:
        for course, avg in records:
            print(f"{course} : {round(avg, 2)}")

    conn.close()

#view database

def view_database():
    conn = sqlite3.connect("student_management.db")
    cursor = conn.cursor()

    print("\n----- STUDENTS -----")
    cursor.execute("SELECT * FROM Students")
    for row in cursor.fetchall():
        print(row)

    print("\n----- COURSES -----")
    cursor.execute("SELECT * FROM Courses")
    for row in cursor.fetchall():
        print(row)

    print("\n----- ENROLLMENTS -----")
    cursor.execute("SELECT * FROM Enrollments")
    for row in cursor.fetchall():
        print(row)

    conn.close()

# ================= EXCEL REPORT =================

def generate_report():
    conn = sqlite3.connect("student_management.db")
    cursor = conn.cursor()

    wb = Workbook()

    # Students Sheet
    ws1 = wb.active
    ws1.title = "Students"

    ws1.append(["ID", "Name", "Age", "Email"])

    cursor.execute("SELECT * FROM Students")
    for row in cursor.fetchall():
        ws1.append(row)

    # Courses Sheet
    ws2 = wb.create_sheet("Courses")
    ws2.append(["ID", "Course", "Duration"])

    cursor.execute("SELECT * FROM Courses")
    for row in cursor.fetchall():
        ws2.append(row)

    # Enrollments Sheet
    ws3 = wb.create_sheet("Enrollments")
    ws3.append(["Enrollment ID", "Student", "Course", "Marks"])

    cursor.execute("""
    SELECT
    Enrollments.enrollment_id,
    Students.student_name,
    Courses.course_name,
    Enrollments.marks
    FROM Enrollments
    JOIN Students
    ON Students.student_id = Enrollments.student_id
    JOIN Courses
    ON Courses.course_id = Enrollments.course_id
    """)

    enrollments = cursor.fetchall()

    for row in enrollments:
        ws3.append(row)

    # Summary Sheet
    ws4 = wb.create_sheet("Summary")

    cursor.execute("SELECT COUNT(*) FROM Students")
    total_students = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM Courses")
    total_courses = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM Enrollments")
    total_enrollments = cursor.fetchone()[0]

    cursor.execute("SELECT marks FROM Enrollments")
    marks = [row[0] for row in cursor.fetchall()]

    highest = 0
    lowest = 0
    average = 0

    if marks:
        highest = marks[0]
        lowest = marks[0]
        total = 0

        for mark in marks:
            if mark > highest:
                highest = mark

            if mark < lowest:
                lowest = mark

            total += mark

        average = total / len(marks)

    cursor.execute("""
    SELECT
    Students.student_name,
    AVG(Enrollments.marks)
    FROM Students
    JOIN Enrollments
    ON Students.student_id = Enrollments.student_id
    GROUP BY Students.student_id
    """)

    records = cursor.fetchall()

    top_name = "N/A"
    highest_avg = -1

    for name, avg in records:
        if avg > highest_avg:
            highest_avg = avg
            top_name = name

    ws4.append(["Total Students", total_students])
    ws4.append(["Total Courses", total_courses])
    ws4.append(["Total Enrollments", total_enrollments])
    ws4.append(["Top Student", top_name])
    ws4.append(["Highest Marks", highest])
    ws4.append(["Lowest Marks", lowest])
    ws4.append(["Average Marks", round(average, 2)])

    wb.save("student_course_report.xlsx")
    conn.close()

    print("Excel Report Generated Successfully.")


# ================= MAIN MENU =================

def main():
    create_database()

    while True:
        print("\n===== STUDENT COURSE MANAGEMENT SYSTEM =====")
        print("1. Add Student")
        print("2. View Students")
        print("3. Add Course")
        print("4. View Courses")
        print("5. Enroll Student Into Course")
        print("6. View Student Enrollments")
        print("7. Update Student Marks")
        print("8. Delete Enrollment")
        print("9. Show Top Performing Student")
        print("10. Show Course Average Marks")
        print("11. Generate Report")
        print("12. Exit")
        print("13. View Database ")

        choice = input("Enter Choice: ")

        if choice == "1":
            add_student()

        elif choice == "2":
            view_students()

        elif choice == "3":
            add_course()

        elif choice == "4":
            view_courses()

        elif choice == "5":
            enroll_student()

        elif choice == "6":
            view_enrollments()

        elif choice == "7":
            update_marks()

        elif choice == "8":
            delete_enrollment()

        elif choice == "9":
            top_student()

        elif choice == "10":
            course_average()

        elif choice == "11":
            generate_report()

        elif choice == "12":
            print("Thank You!")
            break
        
        elif choice == "13":
            view_database()
            

        else:
            print("Invalid Choice.")


main()